# Standard library imports
from openpyxl import load_workbook
import re
from collections import defaultdict
from pprint import pprint


def getCourseStringMatchSet(course_names: str, pattern: str) -> list:
    # pattern = "[456]HSK[0-9]{4}"
    try:
        result = re.findall(pattern, course_names)
    except:
        result = []
    return set(result)


if __name__ == "__main__":
    wb = load_workbook(
        filename="/home/mikel/dev/AdventOfCode/python/2021/ClairesDataTweak/Courses.xlsx"
    )
    sheet_ranges = wb["Level 4"]

    course_to_knowledge = defaultdict(set)
    course_to_skill = defaultdict(set)

    # print(type(full_list))
    for learning_objective in range(7, 58):
        # knowledges = defaultdict(set)
        # skills = defaultdict(set)
        courses = getCourseStringMatchSet(
            sheet_ranges.cell(learning_objective, 2).value, "[456]HSK[0-9]{4}"
        )
        print("type of courses")
        print(type(courses))
        if not courses:
            continue
        knowledges = getCourseStringMatchSet(
            sheet_ranges.cell(learning_objective, 3).value, "K[0-9]{1,}"
        )
        print("type of knowledges")
        print(type(knowledges))
        skills = getCourseStringMatchSet(
            sheet_ranges.cell(learning_objective, 4).value, "S[0-9]{1,}"
        )

        for course in courses:
            if course in course_to_knowledge:
                course_to_knowledge[course] = course_to_knowledge[course].union(
                    knowledges
                )
            else:
                course_to_knowledge[course] = knowledges
            if course in course_to_skill:
                course_to_skill[course] = course_to_skill[course].union(skills)
            else:
                course_to_skill[course] = skills

    wb.create_sheet("NewForClaire")
    clairesheet = wb["NewForClaire"]
    coursecounter = 1
    for course in sorted(course_to_knowledge):
        cellref = clairesheet.cell(coursecounter, 1)
        cellref.value = course
        cellref = clairesheet.cell(coursecounter, 2)
        cellref.value = ", ".join(sorted(course_to_knowledge[course]))
        cellref = clairesheet.cell(coursecounter, 3)
        cellref.value = ", ".join(sorted(course_to_skill[course]))
        coursecounter += 1
    wb.save("New.xlsx")
